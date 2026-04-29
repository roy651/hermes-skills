"""
XLS export using openpyxl.

Generates two files:
  assignments.xlsx — one row per navigation task
  pairings.xlsx    — one row per pair
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SI_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_IF_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_RTL = Alignment(horizontal="right", vertical="center", readingOrder=2)
_CENTER = Alignment(horizontal="center", vertical="center")


def _header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for ci, cell in enumerate(ws[ws.max_row], start=1):
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _RTL


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _rtl_sheet(ws) -> None:
    """Enable RTL display for the sheet."""
    ws.sheet_view.rightToLeft = True


# ---------------------------------------------------------------------------
# Points export
# ---------------------------------------------------------------------------

def export_points(points_db: list[dict], output_dir: str) -> str:
    """Write points.xlsx. Returns file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "נקודות ניווט"
    _rtl_sheet(ws)
    _header_row(ws, ["מזהה", "X (מזרח)", "Y (צפון)", "תיאור"])
    for p in sorted(points_db, key=lambda x: x["id"]):
        ws.append([p["id"], p["x"], p["y"], p.get("description", "")])
        for cell in ws[ws.max_row]:
            cell.alignment = _RTL
    _auto_width(ws)
    out_path = Path(output_dir) / "points.xlsx"
    wb.save(str(out_path))
    return str(out_path)


# ---------------------------------------------------------------------------
# Assignments export
# ---------------------------------------------------------------------------

def export_assignments(
    assignments: list[dict],
    points_db: list[dict],
    output_dir: str,
) -> str:
    """Write assignments.xlsx. Returns file path."""
    pt_map = {p["id"]: p for p in points_db}
    n_pts = max((len(a["points"]) for a in assignments), default=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "משימות ניווט"
    _rtl_sheet(ws)

    headers = ["מס' משימה", "מקטע"] + [f"נקודה {i+1}" for i in range(n_pts)] + ["אורך (ק\"מ)"]
    _header_row(ws, headers)

    for a in sorted(assignments, key=lambda x: x["index"]):
        fill = _SI_FILL if "נה→" in a["section"] else _IF_FILL
        pts_cells = []
        for pid in a["points"]:
            pt = pt_map.get(pid)
            label = f"{pid}" + (f" – {pt['description']}" if pt and pt.get("description") else "")
            pts_cells.append(label)
        # Pad to n_pts
        pts_cells += [""] * (n_pts - len(pts_cells))

        row_data = [a["index"], a["section"]] + pts_cells + [a["length_km"]]
        ws.append(row_data)

        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = _RTL

    _auto_width(ws)

    out_path = Path(output_dir) / "assignments.xlsx"
    wb.save(str(out_path))
    return str(out_path)


# ---------------------------------------------------------------------------
# Pairings export
# ---------------------------------------------------------------------------

def export_pairings(
    pairings: list[dict],
    output_dir: str,
) -> str:
    """Write pairings.xlsx. Returns file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "שיבוץ זוגות"
    _rtl_sheet(ws)

    headers = [
        "מס' זוג",
        "משתתף א", "ציון א", "מס' משימה א", "מקטע א", "אורך א (ק\"מ)",
        "משתתף ב", "ציון ב", "מס' משימה ב", "מקטע ב", "אורך ב (ק\"מ)",
    ]
    _header_row(ws, headers)

    for pr in sorted(pairings, key=lambda x: x["pair_index"]):
        row_data = [
            pr["pair_index"],
            pr["p1_name"],
            round(pr["p1_score"], 1),
            pr["p1_task_index"] or "",
            pr["p1_section"],
            pr["p1_length_km"],
            pr["p2_name"],
            round(pr["p2_score"], 1) if pr["p2_name"] else "",
            pr["p2_task_index"] or "",
            pr["p2_section"],
            pr["p2_length_km"] if pr["p2_name"] else "",
        ]
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.alignment = _RTL

    _auto_width(ws)

    out_path = Path(output_dir) / "pairings.xlsx"
    wb.save(str(out_path))
    return str(out_path)


# ---------------------------------------------------------------------------
# Combined export
# ---------------------------------------------------------------------------

def export_solo_a(
    solo_assignments: list[dict],
    points_db: list[dict],
    output_dir: str,
) -> str:
    """Write solo_assignments.xlsx for solo-A mode (one row per participant). Returns file path."""
    pt_map = {p["id"]: p for p in points_db}
    n_pts = max((len(a["points"]) for a in solo_assignments), default=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "שיבוץ יחידני"
    _rtl_sheet(ws)

    headers = (
        ["מס'", "שם", "ציון", "מס' משימה", "מקטע"]
        + [f"נקודה {i + 1}" for i in range(n_pts)]
        + ["אורך (ק\"מ)"]
    )
    _header_row(ws, headers)

    for a in sorted(solo_assignments, key=lambda x: x["index"]):
        pts_cells = []
        for pid in a["points"]:
            pt = pt_map.get(pid)
            pts_cells.append(f"{pid}" + (f" – {pt['description']}" if pt and pt.get("description") else ""))
        pts_cells += [""] * (n_pts - len(pts_cells))

        row = [a["index"], a["name"], round(a["score"], 1), a["task_index"], a["section"]] + pts_cells + [a["length_km"]]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = _RTL

    _auto_width(ws)
    out_path = Path(output_dir) / "solo_assignments.xlsx"
    wb.save(str(out_path))
    return str(out_path)


def export_solo_mid(
    solo_assignments: list[dict],
    points_db: list[dict],
    output_dir: str,
) -> str:
    """Write solo_mid_assignments.xlsx for solo-mid mode (one row per participant, two tasks). Returns file path."""
    pt_map = {p["id"]: p for p in points_db}
    n_si = max((len(a["si_points"]) for a in solo_assignments), default=0)
    n_if = max((len(a["if_points"]) for a in solo_assignments), default=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "שיבוץ יחידני עם ביניים"
    _rtl_sheet(ws)

    headers = (
        ["מס'", "שם", "ציון"]
        + ["מס' משימה נה→נב"] + [f"נקודה ס{i + 1}" for i in range(n_si)] + ["אורך נה→נב (ק\"מ)"]
        + ["מס' משימה נב→נס"] + [f"נקודה ב{i + 1}" for i in range(n_if)] + ["אורך נב→נס (ק\"מ)"]
    )
    _header_row(ws, headers)

    def _pt_labels(pids: list, n: int) -> list:
        cells = []
        for pid in pids:
            pt = pt_map.get(pid)
            cells.append(f"{pid}" + (f" – {pt['description']}" if pt and pt.get("description") else ""))
        return cells + [""] * (n - len(cells))

    for a in sorted(solo_assignments, key=lambda x: x["index"]):
        row = (
            [a["index"], a["name"], round(a["score"], 1)]
            + [a["si_task_index"]] + _pt_labels(a["si_points"], n_si) + [a["si_length_km"]]
            + [a["if_task_index"]] + _pt_labels(a["if_points"], n_if) + [a["if_length_km"]]
        )
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = _RTL

    _auto_width(ws)
    out_path = Path(output_dir) / "solo_mid_assignments.xlsx"
    wb.save(str(out_path))
    return str(out_path)


def export_combined(
    pairings: list[dict],
    points_db: list[dict],
    output_dir: str,
) -> str:
    """Write combined.xlsx with one row per pair: SI navigator then IF navigator.

    Columns: pair index | SI name | SI pt1..N | IF name | IF pt1..N
    SI = start-to-intermediate (neh->nav), IF = intermediate-to-finish (nav->nes).
    """
    pt_map = {p["id"]: p for p in points_db}

    def _pt_label(pid: int) -> str:
        pt = pt_map.get(pid)
        return f"{pid}" + (f" - {pt['description']}" if pt and pt.get("description") else "")

    def _is_si(section: str) -> bool:
        return "\u05e0\u05d4" in section  # neh (start)

    max_si = max(
        (len(pr["p1_points"]) if _is_si(pr["p1_section"]) else len(pr["p2_points"]) for pr in pairings),
        default=0,
    )
    max_if = max(
        (len(pr["p2_points"]) if _is_si(pr["p1_section"]) else len(pr["p1_points"]) for pr in pairings),
        default=0,
    )

    si_hdr = "\u05e9\u05dd \u05d0 (\u05e0\u05d4\u2192\u05e0\u05d1)"
    if_hdr = "\u05e9\u05dd \u05d1 (\u05e0\u05d1\u2192\u05e0\u05e1)"
    headers = (
        ["\u05de\u05e1 \u05d6\u05d5\u05d2"]
        + [si_hdr] + [f"\u05e0\u05e7 \u05d0{i+1}" for i in range(max_si)]
        + [if_hdr] + [f"\u05e0\u05e7 \u05d1{i+1}" for i in range(max_if)]
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "\u05e9\u05d9\u05d1\u05d5\u05e5 \u05de\u05e9\u05d5\u05dc\u05d1"
    _rtl_sheet(ws)
    _header_row(ws, headers)

    for pr in sorted(pairings, key=lambda x: x["pair_index"]):
        p1_is_si = _is_si(pr["p1_section"])
        si_name = pr["p1_name"] if p1_is_si else pr["p2_name"]
        si_pts  = pr["p1_points"] if p1_is_si else pr["p2_points"]
        if_name = pr["p2_name"] if p1_is_si else pr["p1_name"]
        if_pts  = pr["p2_points"] if p1_is_si else pr["p1_points"]

        si_cells = [_pt_label(pid) for pid in si_pts] + [""] * (max_si - len(si_pts))
        if_cells = [_pt_label(pid) for pid in if_pts] + [""] * (max_if - len(if_pts))

        ws.append([pr["pair_index"]] + [si_name] + si_cells + [if_name] + if_cells)
        for cell in ws[ws.max_row]:
            cell.alignment = _RTL

    _auto_width(ws)
    out_path = Path(output_dir) / "combined.xlsx"
    wb.save(str(out_path))
    return str(out_path)
